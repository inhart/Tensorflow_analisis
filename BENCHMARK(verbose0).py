# %% [markdown]
# ## 0. Objetivo

# %% [markdown]
# ### Analisis del mismo modelo sobre los mismos datos pero con distintas estrategias de compilacion del modelo con la libreria Tensorflow
#

# %% [markdown]
# ## Arquitectura, version Cuda...

# %%
# nvidia-smi

# %%
import os

os.environ["CUDA_VISIBLE_DEVICES"] = "0"

# %% [markdown]
# Seguidamente establecemos una strategia global de espejp con la gpu y una reserva que haremos en la ram

# %%
import tensorflow as tf

tf.debugging.set_log_device_placement(False)
gpus = tf.config.list_logical_devices("GPU")
strategy0 = tf.distribute.MirroredStrategy(gpus)
strategy1 = tf.distribute.MirroredStrategy(
    cross_device_ops=tf.distribute.NcclAllReduce()
)
strategy2 = tf.distribute.MirroredStrategy(
    devices=gpus.append(
        [tf.config.experimental.VirtualDeviceConfiguration(memory_limit=2048)]
    ),
    cross_device_ops=tf.distribute.ReductionToOneDevice(),
)

tf.config.list_physical_devices("GPU")
tf.test.is_built_with_cuda()

# %% [markdown]
# Comprobamos

# %%
gpus = tf.config.list_physical_devices("GPU")
if gpus:
    # Restrict TensorFlow to only use the first GPU
    try:
        tf.config.set_visible_devices(gpus[0], "GPU")
        logical_gpus = tf.config.list_logical_devices("GPU")
        print(len(gpus), "Physical GPUs,", len(logical_gpus), "Logical GPU")
    except RuntimeError as e:
        # Visible devices must be set before GPUs have been initialized
        print(e)


# %%
# Importar bibliotecas necesarias

import os
from openpyxl import load_workbook

import numpy as np  # Para trabajar con arreglos numéricos y generar datos sintéticos
import matplotlib.pyplot as plt  # Para graficar resultados
import seaborn as sns  # Para crear gráficos más atractivos
from sklearn.metrics import (
    confusion_matrix,
)  # Para analizar errores con una matriz de confusión
from sklearn.model_selection import (
    train_test_split,
)  # Para dividir los datos en entrenamiento y validación

import pandas as pd

# Estas son herramientas para construir, entrenar y evaluar redes neuronales.
import keras as keras  # Biblioteca para trabajar con redes neuronales
import tensorflow as tf

from keras.models import Sequential
from keras.layers import Dense, InputLayer, Dropout, BatchNormalization, LeakyReLU
from keras.optimizers import Adam
from keras.callbacks import EarlyStopping, ReduceLROnPlateau, ModelCheckpoint
from keras.regularizers import l2


# ------------------------------------------------
import sklearn

from sklearn.preprocessing import (
    StandardScaler,
    MinMaxScaler,
    RobustScaler,
)  # Para escalar los datos
from sklearn.metrics import mean_squared_error, mean_absolute_error, r2_score
import random
from time import time

# %%

# Fijar semillas para reproducibilidad
SEED = 5785630  # random.randint(0,100000)	  # Generar un número aleatorio entre 0 y 1000
np.random.seed(SEED)
tf.random.set_seed(SEED)
random.seed(SEED)
sklearn.random.seed(SEED)
print(f"Seed: {SEED}")  # Imprimir la semilla utilizada


# %%
def reparar_nombres(file):
    import os

    if " " in file:
        os.rename(file, file.replace(" ", "_"))
    return file.replace(" ", "_")


scores = {}

# %%
# pillar los xls de la carpeta y los mete en un dataframe conjunto

path = "./"
archivos = [x for x in os.listdir(path) if x.endswith(".xlsx")]
df = pd.DataFrame()
df_cols = ""
for archivo in archivos:
    archivo = reparar_nombres(archivo)

    archivo = path + str(archivo)

    xlsx = load_workbook(archivo)

    paginas = xlsx.sheetnames

    dfx = pd.read_excel(archivo, sheet_name=paginas[0])

    dfy = pd.read_excel(archivo, sheet_name=paginas[1])

    df_tmp = pd.concat([dfx, dfy], axis=1)
    if len(df.columns) != 0:
        df.columns = df_tmp.columns
    df = pd.concat([df, df_tmp], axis=0, ignore_index=True)

df = df.dropna(axis=0, how="any")
df = df.drop_duplicates()
df.to_csv("df.csv", index=False)


# %%
df = pd.read_csv("df.csv")

# %%
dfx = df.iloc[:, :-8]
dfy = df.iloc[:, -8:]

# %%
import re

for i in dfy.columns:
    dfy[i] = dfy[i].apply(
        lambda x: x if type(x) is not str else float(re.sub(r"[^0-9.]", "", x))
    )

for i in dfx.columns:
    dfx[i] = dfx[i].apply(
        lambda x: x if type(x) is not str else float(re.sub(r"[^0-9.]", "", x))
    )

dfx = dfx.astype(float)
dfy = dfy.astype(float)

# %%
dfx.describe()

# %%
dfy.describe()


# %%

N_Datos = dfx.shape[0]  # Se generarán 81 ejemplos o muestras.
N_Dimensiones_entrada_X = dfx.shape[
    1
]  # Cada muestra tendrá 800 características (dimensiones de entrada).
N_Dimensiones_salida_Y = dfy.shape[1]  # Habrá 8 clases posibles para la salida.

X = pd.DataFrame(
    dfx
)  # Genera datos de entrada (X) de forma aleatoria con valores flotantes entre 0 y 1.
y = pd.DataFrame(
    dfy
)  # Genera etiquetas de salida (Y) de forma aleatoria con valores enteros entre 0 y 9.


# %%


# --- Preprocesamiento ---
scaler_x = StandardScaler()
X_scaled = scaler_x.fit_transform(X)
scaler_y = StandardScaler()
y_scaled = scaler_y.fit_transform(y)

X_train, X_test, Y_train_scaled, Y_test_scaled = train_test_split(
    X_scaled, y_scaled, test_size=0.2, random_state=SEED
)


# %%
["mse" for _ in range(0, 8)]

# %%
import tensorflow as tf

print("Versión de TensorFlow:", tf.__version__)
print("GPUs detectadas:", tf.config.list_physical_devices("GPU"))


# %% [markdown]
# ## GPU0

# %%

# --- Red neuronal regresora multisalida ---
start_time = time()
early = EarlyStopping(monitor="loss", patience=25, restore_best_weights=True)
reduce_lr = ReduceLROnPlateau(monitor="loss", patience=15, factor=0.5)
with tf.device("/device:GPU:0"):
    # Definir el modelo
    model0 = Sequential(
        [
            InputLayer(
                input_shape=(N_Dimensiones_entrada_X,)
            ),  # Capa de entrada con 8000 características
            BatchNormalization(),  # Normalización por lotes para mejorar la convergencia
            LeakyReLU(0.5),
            Dense(
                512, activation="leaky_relu", kernel_regularizer=l2(0.001)
            ),  # Capa oculta con 1024 neuronas
            Dropout(0.2),  # Capa de abandono para evitar el sobreajuste
            LeakyReLU(0.25),
            Dense(128, activation="leaky_relu"),  # Otra capa oculta con 512 neuronas
            Dropout(0.1),  # Capa de abandono para evitar el sobreajuste
            Dense(
                16, activation="relu", kernel_regularizer=l2(0.001)
            ),  # Otra capa oculta con 256 neuronas
            Dense(N_Dimensiones_salida_Y),  # Capa de salida con 8 valores (Target)
        ]
    )

    adam = Adam(learning_rate=0.001)
    # Compilar el modelo
    model0.compile(
        optimizer=adam,
        loss=["mse", "mse", "mse", "mse", "mse", "mse", "mse", "mse"],
        loss_weights=(0.9, 0.4, 0.3, 0.4, 0.8, 0.1, 0.3, 0.7),
        metrics=["mae", "mse"],
    )  # Función de pérdida MSE y métricas MAE y MSE

end_time = time()
print(f"Tiempo de compilacion del modelo 0: {end_time - start_time:.2f} segundos")
# Resumen del modelo

# %%

model0.summary()
start_time = time()
# --- Entrenamiento ---
history0 = model0.fit(
    X_train,
    Y_train_scaled,
    epochs=500,
    batch_size=32,
    validation_data=(X_test, Y_test_scaled),
    verbose=0,
    callbacks=[early, reduce_lr],
)
stop_time = time()
print(f"Tiempo de entrenamiento del modelo 0: {stop_time - start_time:.2f} segundos")


# %%


Y_pred_scaled0 = model0.predict(X_test)

# --- Evaluación ---
Y_pred0 = scaler_y.inverse_transform(Y_pred_scaled0)
Y_test0 = scaler_y.inverse_transform(Y_test_scaled)
aux = {}
aux["r2"] = []
aux["mae"] = []
aux["mse"] = []
aux["rmse"] = []


""" # --- Evaluación ---
# Errores por cada salida
for i in range(N_Dimensiones_salida_Y):
    print(f"--- Output {i+1} ---")
    print("  R2 :", r2_score(Y_test0[:, i], Y_pred0[:, i]))
    aux['r2'].append(r2_score(Y_test0[:, i], Y_pred0[:, i]))
    print("MAE:", mean_absolute_error(Y_test0[:, i], Y_pred0[:, i]))
    aux['mae'].append(mean_absolute_error(Y_test0[:, i], Y_pred0[:, i]))
    print("  MSE:", mean_squared_error(Y_test0[:, i], Y_pred0[:, i]))
    aux['mse'].append(mean_squared_error(Y_test0[:, i], Y_pred0[:, i]))
    print("  RMSE:", np.sqrt(mean_squared_error(Y_test0[:, i], Y_pred0[:, i])))
    aux['rmse'].append(np.sqrt(mean_squared_error(Y_test0[:, i], Y_pred0[:, i])))
aux = pd.DataFrame(aux)

import matplotlib.pyplot as plt

for i in range(N_Dimensiones_salida_Y):
    plt.figure(figsize=(5,4))
    plt.scatter(Y_test0[:, i], Y_pred0[:, i], alpha=0.5)
    plt.plot([Y_test0[:, i].min(), Y_test0[:, i].max()],
             [Y_test0[:, i].min(), Y_test0[:, i].max()],
             'r--')
    plt.xlabel("Valor real")
    plt.ylabel("Predicción")
    plt.title(f"{dfy.keys()[i]} - Pred vs Real")
    plt.grid()
    plt.tight_layout()
    plt.show()
     """
mae_total0 = mean_absolute_error(Y_test0.flatten(), Y_pred0.flatten())
mse_total0 = mean_squared_error(Y_test0.flatten(), Y_pred0.flatten())
r2_total0 = r2_score(Y_test0, Y_pred0)

print("\n--- Total ---")
print("R² :", r2_total0)
print("MAE:", mae_total0)
print("MSE:", mse_total0)

# %% [markdown]
# ## tf.distribute.MirroredStrategy(gpus)

# %%

# --- Red neuronal regresora multisalida ---
start_time = time()
early = EarlyStopping(monitor="loss", patience=10, restore_best_weights=True)
reduce_lr = ReduceLROnPlateau(monitor="loss", patience=5, factor=0.5)
with strategy0.scope():
    # Definir el modelo
    model1 = Sequential(
        [
            InputLayer(
                input_shape=(N_Dimensiones_entrada_X,)
            ),  # Capa de entrada con 8000 características
            BatchNormalization(),  # Normalización por lotes para mejorar la convergencia
            LeakyReLU(0.5),
            Dense(
                512, activation="leaky_relu", kernel_regularizer=l2(0.001)
            ),  # Capa oculta con 1024 neuronas
            Dropout(0.2),  # Capa de abandono para evitar el sobreajuste
            LeakyReLU(0.25),
            Dense(128, activation="leaky_relu"),  # Otra capa oculta con 512 neuronas
            Dropout(0.1),  # Capa de abandono para evitar el sobreajuste
            Dense(
                16, activation="relu", kernel_regularizer=l2(0.001)
            ),  # Otra capa oculta con 256 neuronas
            Dense(N_Dimensiones_salida_Y),  # Capa de salida con 8 valores (Target)
        ]
    )

    adam = Adam(learning_rate=0.001)
    # Compilar el modelo
    model1.compile(
        optimizer=adam,
        loss=["mse", "mse", "mse", "mse", "mse", "mse", "mse", "mse"],
        loss_weights=(0.1, 1, 1, 1, 1, 1, 1, 1),
        metrics=["mae", "mse"],
    )  # Función de pérdida MSE y métricas MAE y MSE
end_time = time()
print(f"Tiempo de compilacion del modelo 1: {end_time - start_time:.2f} segundos")

# %%

# Resumen del modelo
model1.summary()
start_time = time()
# --- Entrenamiento ---
history1 = model1.fit(
    X_train,
    Y_train_scaled,
    epochs=500,
    batch_size=32,
    validation_data=(X_test, Y_test_scaled),
    verbose=0,
    callbacks=[early, reduce_lr],
)
stop_time = time()
print(f"Tiempo de entrenamiento del modelo 1: {stop_time - start_time:.2f} segundos")

# %%


Y_pred_scaled1 = model1.predict(X_test)

# --- Evaluación ---
Y_pred1 = scaler_y.inverse_transform(Y_pred_scaled1)
Y_test1 = scaler_y.inverse_transform(Y_test_scaled)
aux = {}
aux["r2"] = []
aux["mae"] = []
aux["mse"] = []
aux["rmse"] = []


""" # --- Evaluación ---
# Errores por cada salida
for i in range(N_Dimensiones_salida_Y):
    print(f"--- Output {i+1} ---")
    print("  R2 :", r2_score(Y_test1[:, i], Y_pred1[:, i]))
    aux['r2'].append(r2_score(Y_test1[:, i], Y_pred1[:, i]))
    print("MAE:", mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mae'].append(mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  MSE:", mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mse'].append(mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  RMSE:", np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
    aux['rmse'].append(np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
aux = pd.DataFrame(aux)

import matplotlib.pyplot as plt

for i in range(N_Dimensiones_salida_Y):
    plt.figure(figsize=(5,4))
    plt.scatter(Y_test1[:, i], Y_pred1[:, i], alpha=0.5)
    plt.plot([Y_test1[:, i].min(), Y_test1[:, i].max()],
             [Y_test1[:, i].min(), Y_test1[:, i].max()],
             'r--')
    plt.xlabel("Valor real")
    plt.ylabel("Predicción")
    plt.title(f"{dfy.keys()[i]} - Pred vs Real")
    plt.grid()
    plt.tight_layout()
    plt.show()
     """
mae_total1 = mean_absolute_error(Y_test1.flatten(), Y_pred1.flatten())
mse_total1 = mean_squared_error(Y_test1.flatten(), Y_pred1.flatten())
r2_total1 = r2_score(Y_test1, Y_pred1)

print("\n--- Total ---")
print("R² :", r2_total1)
print("MAE:", mae_total1)
print("MSE:", mse_total1)


# %% [markdown]
# ## tf.distribute.MirroredStrategy(cross_device_ops=tf.distribute.NcclAllReduce())

# %%

# --- Red neuronal regresora multisalida ---
start_time = time()
early = EarlyStopping(monitor="loss", patience=10, restore_best_weights=True)
reduce_lr = ReduceLROnPlateau(monitor="loss", patience=5, factor=0.5)
with strategy1.scope():
    # Definir el modelo
    model1 = Sequential(
        [
            InputLayer(
                input_shape=(N_Dimensiones_entrada_X,)
            ),  # Capa de entrada con 8000 características
            BatchNormalization(),  # Normalización por lotes para mejorar la convergencia
            LeakyReLU(0.5),
            Dense(
                512, activation="leaky_relu", kernel_regularizer=l2(0.001)
            ),  # Capa oculta con 1024 neuronas
            Dropout(0.2),  # Capa de abandono para evitar el sobreajuste
            LeakyReLU(0.25),
            Dense(128, activation="leaky_relu"),  # Otra capa oculta con 512 neuronas
            Dropout(0.1),  # Capa de abandono para evitar el sobreajuste
            Dense(
                16, activation="relu", kernel_regularizer=l2(0.001)
            ),  # Otra capa oculta con 256 neuronas
            Dense(N_Dimensiones_salida_Y),  # Capa de salida con 8 valores (Target)
        ]
    )

    adam = Adam(learning_rate=0.001)
    # Compilar el modelo
    model1.compile(
        optimizer=adam,
        loss=["mse", "mse", "mse", "mse", "mse", "mse", "mse", "mse"],
        loss_weights=(0.1, 1, 1, 1, 1, 1, 1, 1),
        metrics=["mae", "mse"],
    )  # Función de pérdida MSE y métricas MAE y MSE
end_time = time()
print(f"Tiempo de compilacion del modelo 1: {end_time - start_time:.2f} segundos")

# %%

# Resumen del modelo
model1.summary()
start_time = time()
# --- Entrenamiento ---
history1 = model1.fit(
    X_train,
    Y_train_scaled,
    epochs=500,
    batch_size=32,
    validation_data=(X_test, Y_test_scaled),
    verbose=0,
    callbacks=[early, reduce_lr],
)
stop_time = time()
print(f"Tiempo de entrenamiento del modelo 1: {stop_time - start_time:.2f} segundos")

# %%


Y_pred_scaled1 = model1.predict(X_test)

# --- Evaluación ---
Y_pred1 = scaler_y.inverse_transform(Y_pred_scaled1)
Y_test1 = scaler_y.inverse_transform(Y_test_scaled)
aux = {}
aux["r2"] = []
aux["mae"] = []
aux["mse"] = []
aux["rmse"] = []


""" # --- Evaluación ---
# Errores por cada salida
for i in range(N_Dimensiones_salida_Y):
    print(f"--- Output {i+1} ---")
    print("  R2 :", r2_score(Y_test1[:, i], Y_pred1[:, i]))
    aux['r2'].append(r2_score(Y_test1[:, i], Y_pred1[:, i]))
    print("MAE:", mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mae'].append(mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  MSE:", mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mse'].append(mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  RMSE:", np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
    aux['rmse'].append(np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
aux = pd.DataFrame(aux)

import matplotlib.pyplot as plt

for i in range(N_Dimensiones_salida_Y):
    plt.figure(figsize=(5,4))
    plt.scatter(Y_test1[:, i], Y_pred1[:, i], alpha=0.5)
    plt.plot([Y_test1[:, i].min(), Y_test1[:, i].max()],
             [Y_test1[:, i].min(), Y_test1[:, i].max()],
             'r--')
    plt.xlabel("Valor real")
    plt.ylabel("Predicción")
    plt.title(f"{dfy.keys()[i]} - Pred vs Real")
    plt.grid()
    plt.tight_layout()
    plt.show() """

mae_total1 = mean_absolute_error(Y_test1.flatten(), Y_pred1.flatten())
mse_total1 = mean_squared_error(Y_test1.flatten(), Y_pred1.flatten())
r2_total1 = r2_score(Y_test1, Y_pred1)

print("\n--- Total ---")
print("R² :", r2_total1)
print("MAE:", mae_total1)
print("MSE:", mse_total1)


# %% [markdown]
# ## tf.distribute.MirroredStrategy(devices=gpus.append([tf.config.experimental.VirtualDeviceConfiguration(memory_limit=2048)]), cross_device_ops=tf.distribute.ReductionToOneDevice(), )

# %%

# --- Red neuronal regresora multisalida ---
start_time = time()
early = EarlyStopping(monitor="loss", patience=10, restore_best_weights=True)
reduce_lr = ReduceLROnPlateau(monitor="loss", patience=5, factor=0.5)
with strategy2.scope():
    # Definir el modelo
    model1 = Sequential(
        [
            InputLayer(
                input_shape=(N_Dimensiones_entrada_X,)
            ),  # Capa de entrada con 8000 características
            BatchNormalization(),  # Normalización por lotes para mejorar la convergencia
            LeakyReLU(0.5),
            Dense(
                512, activation="leaky_relu", kernel_regularizer=l2(0.001)
            ),  # Capa oculta con 1024 neuronas
            Dropout(0.2),  # Capa de abandono para evitar el sobreajuste
            LeakyReLU(0.25),
            Dense(128, activation="leaky_relu"),  # Otra capa oculta con 512 neuronas
            Dropout(0.1),  # Capa de abandono para evitar el sobreajuste
            Dense(
                16, activation="relu", kernel_regularizer=l2(0.001)
            ),  # Otra capa oculta con 256 neuronas
            Dense(N_Dimensiones_salida_Y),  # Capa de salida con 8 valores (Target)
        ]
    )

    adam = Adam(learning_rate=0.001)
    # Compilar el modelo
    model1.compile(
        optimizer=adam,
        loss=["mse", "mse", "mse", "mse", "mse", "mse", "mse", "mse"],
        loss_weights=(0.1, 1, 1, 1, 1, 1, 1, 1),
        metrics=["mae", "mse"],
    )  # Función de pérdida MSE y métricas MAE y MSE
end_time = time()
print(f"Tiempo de compilacion del modelo 1: {end_time - start_time:.2f} segundos")

# %%

# Resumen del modelo
model1.summary()
start_time = time()
# --- Entrenamiento ---
history1 = model1.fit(
    X_train,
    Y_train_scaled,
    epochs=500,
    batch_size=32,
    validation_data=(X_test, Y_test_scaled),
    verbose=0,
    callbacks=[early, reduce_lr],
)
stop_time = time()
print(f"Tiempo de entrenamiento del modelo 1: {stop_time - start_time:.2f} segundos")

# %%


Y_pred_scaled1 = model1.predict(X_test)

# --- Evaluación ---
Y_pred1 = scaler_y.inverse_transform(Y_pred_scaled1)
Y_test1 = scaler_y.inverse_transform(Y_test_scaled)
aux = {}
aux["r2"] = []
aux["mae"] = []
aux["mse"] = []
aux["rmse"] = []


""" # --- Evaluación ---
# Errores por cada salida
for i in range(N_Dimensiones_salida_Y):
    print(f"--- Output {i+1} ---")
    print("  R2 :", r2_score(Y_test1[:, i], Y_pred1[:, i]))
    aux['r2'].append(r2_score(Y_test1[:, i], Y_pred1[:, i]))
    print("MAE:", mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mae'].append(mean_absolute_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  MSE:", mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    aux['mse'].append(mean_squared_error(Y_test1[:, i], Y_pred1[:, i]))
    print("  RMSE:", np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
    aux['rmse'].append(np.sqrt(mean_squared_error(Y_test1[:, i], Y_pred1[:, i])))
aux = pd.DataFrame(aux)

import matplotlib.pyplot as plt

for i in range(N_Dimensiones_salida_Y):
    plt.figure(figsize=(5,4))
    plt.scatter(Y_test1[:, i], Y_pred1[:, i], alpha=0.5)
    plt.plot([Y_test1[:, i].min(), Y_test1[:, i].max()],
             [Y_test1[:, i].min(), Y_test1[:, i].max()],
             'r--')
    plt.xlabel("Valor real")
    plt.ylabel("Predicción")
    plt.title(f"{dfy.keys()[i]} - Pred vs Real")
    plt.grid()
    plt.tight_layout()
    plt.show()
     """
mae_total1 = mean_absolute_error(Y_test1.flatten(), Y_pred1.flatten())
mse_total1 = mean_squared_error(Y_test1.flatten(), Y_pred1.flatten())
r2_total1 = r2_score(Y_test1, Y_pred1)

print("\n--- Total ---")
print("R² :", r2_total1)
print("MAE:", mae_total1)
print("MSE:", mse_total1)


# %% [markdown]
# ## CPU:0

# %%

# --- Red neuronal regresora multisalida ---
start_time = time()
early = EarlyStopping(monitor="loss", patience=10, restore_best_weights=True)
reduce_lr = ReduceLROnPlateau(monitor="loss", patience=5, factor=0.5)
with tf.device("/device:CPU:0"):
    # Definir el modelo
    model2 = Sequential(
        [
            InputLayer(
                input_shape=(N_Dimensiones_entrada_X,)
            ),  # Capa de entrada con 8000 características
            BatchNormalization(),  # Normalización por lotes para mejorar la convergencia
            LeakyReLU(0.5),
            Dense(
                512, activation="leaky_relu", kernel_regularizer=l2(0.001)
            ),  # Capa oculta con 1024 neuronas
            Dropout(0.2),  # Capa de abandono para evitar el sobreajuste
            LeakyReLU(0.25),
            Dense(128, activation="leaky_relu"),  # Otra capa oculta con 512 neuronas
            Dropout(0.1),  # Capa de abandono para evitar el sobreajuste
            Dense(
                16, activation="relu", kernel_regularizer=l2(0.001)
            ),  # Otra capa oculta con 256 neuronas
            Dense(N_Dimensiones_salida_Y),  # Capa de salida con 8 valores (Target)
        ]
    )

    adam = Adam(learning_rate=0.001)
    # Compilar el modelo
    model2.compile(
        optimizer=adam,
        loss=["mse", "mse", "mse", "mse", "mse", "mse", "mse", "mse"],
        loss_weights=(0.1, 1, 1, 1, 1, 1, 1, 1),
        metrics=["mae", "mse"],
    )  # Función de pérdida MSE y métricas MAE y MSE
stop_time = time()
print(f"Tiempo de compilacion del modelo 2: {stop_time - start_time:.2f} segundos")

# %%

# Resumen del modelo
model2.summary()
start_time = time()
# --- Entrenamiento ---
history2 = model2.fit(
    X_train,
    Y_train_scaled,
    epochs=500,
    batch_size=32,
    validation_data=(X_test, Y_test_scaled),
    verbose=0,
    callbacks=[early, reduce_lr],
)
stop_time = time()
print(f"Tiempo de entrenamiento del modelo 2: {stop_time - start_time:.2f} segundos")


# %%


Y_pred_scaled2 = model2.predict(X_test)

# --- Evaluación ---
Y_pred2 = scaler_y.inverse_transform(Y_pred_scaled2)
Y_test2 = scaler_y.inverse_transform(Y_test_scaled)
aux = {}
aux["r2"] = []
aux["mae"] = []
aux["mse"] = []
aux["rmse"] = []

"""
# --- Evaluación ---
# Errores por cada salida
for i in range(N_Dimensiones_salida_Y):
    print(f"--- Output {i+1} ---")
    print("  R2 :", r2_score(Y_test2[:, i], Y_pred2[:, i]))
    aux['r2'].append(r2_score(Y_test2[:, i], Y_pred2[:, i]))
    print("MAE:", mean_absolute_error(Y_test2[:, i], Y_pred2[:, i]))
    aux['mae'].append(mean_absolute_error(Y_test2[:, i], Y_pred2[:, i]))
    print("  MSE:", mean_squared_error(Y_test2[:, i], Y_pred2[:, i]))
    aux['mse'].append(mean_squared_error(Y_test2[:, i], Y_pred2[:, i]))
    print("  RMSE:", np.sqrt(mean_squared_error(Y_test2[:, i], Y_pred2[:, i])))
    aux['rmse'].append(np.sqrt(mean_squared_error(Y_test2[:, i], Y_pred2[:, i])))
aux = pd.DataFrame(aux)

import matplotlib.pyplot as plt

for i in range(N_Dimensiones_salida_Y):
    plt.figure(figsize=(5,4))
    plt.scatter(Y_test2[:, i], Y_pred2[:, i], alpha=0.5)
    plt.plot([Y_test2[:, i].min(), Y_test2[:, i].max()],
            [Y_test2[:, i].min(), Y_test2[:, i].max()],
            'r--')
    plt.xlabel("Valor real")
    plt.ylabel("Predicción")
    plt.title(f"{dfy.keys()[i]} - Pred vs Real")
    plt.grid()
    plt.tight_layout()
    plt.show()
"""
mae_total2 = mean_absolute_error(Y_test2.flatten(), Y_pred2.flatten())
mse_total2 = mean_squared_error(Y_test2.flatten(), Y_pred2.flatten())
r2_total2 = r2_score(Y_test2, Y_pred2)

print("\n--- Total ---")
print("R² :", r2_total2)
print("MAE:", mae_total2)
print("MSE:", mse_total2)
